#!/usr/bin/python3
#
#  Copyright (C) 2018-2019 Sustainable Energy Now Inc., Angus King
#
#  powermatch.py - This file is possibly part of SIREN.
#
#  SIREN is free software: you can redistribute it and/or modify
#  it under the terms of the GNU Affero General Public License as
#  published by the Free Software Foundation, either version 3 of
#  the License, or (at your option) any later version.
#
#  SIREN is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU Affero General Public License for more details.
#
#  You should have received a copy of the GNU Affero General
#  Public License along with SIREN.  If not, see
#  <http://www.gnu.org/licenses/>.
#
import os
import sys
import time
from PyQt4 import QtCore, QtGui
import displayobject
import displaytable
from credits import fileVersion
from matplotlib import rcParams
import matplotlib.pyplot as plt
import numpy as np
import openpyxl as oxl
import random
# from openpyxl.utils import get_column_letter
from parents import getParents
from senuser import getUser
from editini import SaveIni
from floaters import ProgressBar
import xlrd
import configparser  # decode .ini file

tech_names = ['Load', 'Onshore Wind', 'Offshore Wind', 'Rooftop PV', 'Fixed PV', 'Single Axis PV',
              'Dual Axis PV', 'Biomass', 'Geothermal', 'Other1', 'CST', 'Shortfall']
col_letters = ' ABCDEFGHIJKLMNOPQRSTUVWXYZ'
C = 0 # same order as self.file_labels
G = 1
O = 2
D = 3
R = 4
def ss_col(col, base=1):
    if base == 1:
        col -= 1
    c1 = col // 26
    c2 = col % 26
    return (col_letters[c1] + col_letters[c2 + 1]).strip()


class ThumbListWidget(QtGui.QListWidget):
    def __init__(self, type, parent=None):
        super(ThumbListWidget, self).__init__(parent)
        self.setIconSize(QtCore.QSize(124, 124))
        self.setDragDropMode(QtGui.QAbstractItemView.DragDrop)
        self.setSelectionMode(QtGui.QAbstractItemView.ExtendedSelection)
        self.setAcceptDrops(True)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.accept()
        else:
            super(ThumbListWidget, self).dragEnterEvent(event)

    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls():
            event.setDropAction(QtCore.Qt.CopyAction)
            event.accept()
        else:
            super(ThumbListWidget, self).dragMoveEvent(event)

    def dropEvent(self, event):
        if event.mimeData().hasUrls():
            event.setDropAction(QtCore.Qt.CopyAction)
            event.accept()
            links = []
            for url in event.mimeData().urls():
                links.append(str(url.toLocalFile()))
            self.emit(QtCore.SIGNAL("dropped"), links)
        else:
            event.setDropAction(QtCore.Qt.MoveAction)
            super(ThumbListWidget, self).dropEvent(event)


class ClickableQLabel(QtGui.QLabel):
    def __init(self, parent):
        QLabel.__init__(self, parent)

    def mousePressEvent(self, event):
        QtGui.QApplication.widgetAt(event.globalPos()).setFocus()
        self.emit(QtCore.SIGNAL('clicked()'))


class Constraint:
    def __init__(self, name, category, capacity_min, capacity_max, rampup_max, rampdown_max,
                 recharge_max, recharge_loss, discharge_max, discharge_loss, parasitic_loss):
        self.name = name
        self.category = category
        try:
            self.capacity_min = float(capacity_min) # minimum run_rate for generator; don't drain below for storage
        except:
            self.capacity_min = 0.
        try:
            self.capacity_max = float(capacity_max) # maximum run_rate for generator; don't drain more than this for storage
        except:
            self.capacity_max = 1.
        try:
            self.recharge_max = float(recharge_max) # can't charge more than this per hour
        except:
            self.recharge_max = 1.
        try:
            self.recharge_loss = float(recharge_loss)
        except:
            self.recharge_loss = 0.
        try:
            self.discharge_max = float(discharge_max) # can't discharge more than this per hour
        except:
            self.discharge_max = 1.
        try:
            self.discharge_loss = float(discharge_loss)
        except:
            self.discharge_loss = 0.
        try:
            self.parasitic_loss = float(parasitic_loss) # daily parasitic loss / hourly ?
        except:
            self.parasitic_loss = 0.
        try:
            self.rampup_max = float(rampup_max)
        except:
            self.rampup_max = 1.
        try:
            self.rampdown_max = float(rampdown_max)
        except:
            self.rampdown_max = 1.


class Facility:
    def __init__(self, name, order, constraint, capacity, lcoe, lcoe_cf, emissions, initial=None):
        self.name = name
        self.constraint = constraint
        try:
            self.order = int(order)
        except:
            self.order = 0.
        try:
            self.capacity = float(capacity)
        except:
            self.capacity = 0.
        try:
            self.lcoe = float(lcoe)
        except:
            self.lcoe = 0.
        try:
            self.lcoe_cf = float(lcoe_cf)
        except:
            self.lcoe_cf = 0.
        try:
            self.emissions = float(emissions)
        except:
            self.emissions = 0.
        try:
            self.initial = float(initial)
        except:
            self.initial = 0.


class Optimisation:
    def __init__(self, name, approach, values): #capacity=None, cap_min=None, cap_max=None, cap_step=None, caps=None):
        self.name = name
        self.approach = approach
        if approach == 'Discrete':
            caps = values.split(' ')
            self.capacities = []
            cap_max = 0.
            for cap in caps:
                try:
                    self.capacities.append(float(cap))
                    cap_max += float(cap)
                except:
                    pass
            self.capacity_min = 0
            self.capacity_max = round(cap_max, 3)
            self.capacity_step = None
        elif approach == 'Range':
            caps = values.split(' ')
            try:
                self.capacity_min = float(caps[0])
            except:
                self.capacity_min = 0
            try:
                self.capacity_max = float(caps[1])
            except:
                self.capacity_max = None
            try:
                self.capacity_step = float(caps[2])
            except:
                self.capacity_step = None
            self.capacities = None
        else:
            self.capacity_min = 0
            self.capacity_max = None
            self.capacity_step = None
            self.capacities = None
        self.capacity = None


class Adjustments(QtGui.QDialog):
    def __init__(self, data, adjustin):
        super(Adjustments, self).__init__()
        self.adjusts = {}
        self.checkbox = {}
        self.labels = {}
        self.results = None
        self.grid = QtGui.QGridLayout()
        self.data = {}
        ctr = 0
        for key, capacity in data:
            if key != 'Load' and capacity is None:
                continue
            self.adjusts[key] = QtGui.QDoubleSpinBox()
            self.adjusts[key].setRange(0, 25)
            self.adjusts[key].setDecimals(3)
            try:
                self.adjusts[key].setValue(adjustin[key])
            except:
                self.adjusts[key].setValue(1.)
            self.data[key] = capacity
            self.adjusts[key].setSingleStep(.1)
            self.adjusts[key].setObjectName(key)
            self.grid.addWidget(QtGui.QLabel(key), ctr, 0)
            self.grid.addWidget(self.adjusts[key], ctr, 1)
            if key != 'Load' or key == 'Load':
                self.adjusts[key].valueChanged.connect(self.adjust)
                self.labels[key] = QtGui.QLabel('')
                self.labels[key].setObjectName(key + 'label')
                if key != 'Load':
                    mw = '{:.0f} MW'.format(capacity * self.adjusts[key].value())
                    self.labels[key].setText(mw)
                self.grid.addWidget(self.labels[key], ctr, 2)
            ctr += 1
        quit = QtGui.QPushButton('Quit', self)
        self.grid.addWidget(quit, ctr, 0)
        quit.clicked.connect(self.quitClicked)
        show = QtGui.QPushButton('Proceed', self)
        self.grid.addWidget(show, ctr, 1)
        show.clicked.connect(self.showClicked)
        frame = QtGui.QFrame()
        frame.setLayout(self.grid)
        self.scroll = QtGui.QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setWidget(frame)
        self.layout = QtGui.QVBoxLayout(self)
        self.layout.addWidget(self.scroll)
        self.setWindowTitle('SIREN - Powermatch - Adjust generators')
        self.setWindowIcon(QtGui.QIcon('sen_icon32.ico'))
        QtGui.QShortcut(QtGui.QKeySequence('q'), self, self.quitClicked)
        self.show()

    def adjust(self):
        key = str(self.sender().objectName())
        mw = '{:.0f} MW'.format(self.data[key] * self.sender().value())
        self.labels[key].setText(mw)

    def closeEvent(self, event):
        event.accept()

    def quitClicked(self):
        self.close()

    def showClicked(self):
        self.results = {}
        for key in list(self.adjusts.keys()):
            self.results[key] = round(self.adjusts[key].value(), 3)
        self.close()

    def getValues(self):
        return self.results

class powerMatch(QtGui.QWidget):

    def __init__(self, help='help.html'):
        super(powerMatch, self).__init__()
        self.help = help
        config = configparser.RawConfigParser()
        if len(sys.argv) > 1:
            config_file = sys.argv[1]
        else:
            config_file = 'SIREN.ini'
        config.read(config_file)
        parents = []
        try:
            parents = getParents(config.items('Parents'))
        except:
            pass
        try:
            base_year = config.get('Base', 'year')
        except:
            base_year = '2012'
        try:
            scenario_prefix = config.get('Files', 'scenario_prefix')
        except:
            scenario_prefix = ''
        try:
            self.scenarios = config.get('Files', 'scenarios')
            if scenario_prefix != '' :
                self.scenarios += '/' + scenario_prefix
            for key, value in parents:
                self.scenarios = self.scenarios.replace(key, value)
            self.scenarios = self.scenarios.replace('$USER$', getUser())
            self.scenarios = self.scenarios.replace('$YEAR$', base_year)
            self.scenarios = self.scenarios[: self.scenarios.rfind('/') + 1]
            if self.scenarios[:3] == '../':
                ups = self.scenarios.split('../')
                me = os.getcwd().split(os.sep)
                me = me[: -(len(ups) - 1)]
                me.append(ups[-1])
                self.scenarios = '/'.join(me)
        except:
            self.scenarios = ''
        self.file_labels = ['Constraints', 'Generators', 'Optimisation', 'Data', 'Results']
        self.ifiles = [''] * 5
        self.isheets = self.file_labels[:]
        del self.isheets[-2:]
        self.adjust_re = False
        self.change_res = True
        self.details = True
        self.carbon_price = 0.
        self.optimise_population = 50
        self.optimise_generations = 20
        self.optimise_stop = 0
        try:
             items = config.items('Powermatch')
             for key, value in items:
                 if key == 'adjust_generators':
                     if value.lower() in ['true', 'on', 'yes']:
                         self.adjust_re = True
                 elif key == 'carbon_price':
                     try:
                         self.carbon_price = float(value)
                     except:
                         pass
                 elif key == 'change_results':
                     if value.lower() in ['false', 'off', 'no']:
                         self.change_res = False
                 elif key[-5:] == '_file':
                     ndx = self.file_labels.index(key[:-5].title())
                     self.ifiles[ndx] = value
                 elif key[-6:] == '_sheet':
                     ndx = self.file_labels.index(key[:-6].title())
                     self.isheets[ndx] = value
                 elif key == 'generators_details':
                     if value.lower() in ['false', 'off', 'no']:
                         self.details = False
                 elif key == 'optimise_population':
                     try:
                         self.optimise_population = int(value)
                     except:
                         pass
                 elif key == 'optimise_generations':
                     try:
                         self.optimise_generations = int(value)
                     except:
                         pass
                 elif key == 'optimise_stop':
                     try:
                         self.optimise_stop = int(value)
                     except:
                         pass
        except:
            pass
        self.opt_progressbar = None
        self.adjustby = None
        self.initUI()

    def initUI(self):
   #     self.tabs = QtGui.QTabWidget()    # Create tabs
   #     tab1 = QtGui.QWidget()
   #     tab2 = QtGui.QWidget()
   #     tab3 = QtGui.QWidget()
   #     tab4 = QtGui.QWidget()
   #     tab5 = QtGui.QWidget()
        self.grid = QtGui.QGridLayout()
        self.files = [None] * 5
        self.sheets = self.file_labels[:]
        del self.sheets[-2:]
        self.updated = False
        edit = [None, None, None]
        r = 0
        for i in range(5):
            self.grid.addWidget(QtGui.QLabel(self.file_labels[i] + ' File:'), r, 0)
            self.files[i] = ClickableQLabel()
            self.files[i].setStyleSheet("background-color: white; border: 1px inset grey; min-height: 22px; border-radius: 4px;")
            self.files[i].setText(self.ifiles[i])
            self.connect(self.files[i], QtCore.SIGNAL('clicked()'), self.fileChanged)
            self.grid.addWidget(self.files[i], r, 1, 1, 3)
            if i < D:
                r += 1
                self.grid.addWidget(QtGui.QLabel(self.file_labels[i] + ' Sheet:'), r, 0)
                self.sheets[i] = QtGui.QComboBox()
                self.sheets[i].addItem(self.isheets[i])
                self.grid.addWidget(self.sheets[i], r, 1, 1, 2)
                edit[i] = QtGui.QPushButton(self.file_labels[i], self)
                self.grid.addWidget(edit[i], r, 3)
                edit[i].clicked.connect(self.editClicked)
            r += 1
        wdth = edit[1].fontMetrics().boundingRect(edit[1].text()).width() + 9
        self.grid.addWidget(QtGui.QLabel('Carbon Price:'), r, 0)
        self.carbon = QtGui.QDoubleSpinBox()
        self.carbon.setRange(0, 200)
        self.carbon.setDecimals(2)
        try:
            self.carbon.setValue(self.carbon_price)
        except:
            self.carbon.setValue(0.)
        self.grid.addWidget(self.carbon, r, 1)
        self.carbon.valueChanged.connect(self.cpchanged)
        self.grid.addWidget(QtGui.QLabel('($/tCO2e. Use only if LCOE excludes carbon price)'), r, 2, 1, 2)
        r += 1
        self.grid.addWidget(QtGui.QLabel('Adjust Generators:'), r, 0)
        self.adjust = QtGui.QCheckBox('(check to adjust/multiply generators capacity data)', self)
        if self.adjust_re:
            self.adjust.setCheckState(QtCore.Qt.Checked)
        self.grid.addWidget(self.adjust, r, 1, 1, 3)
        r += 1
        self.grid.addWidget(QtGui.QLabel('Dispatch Order:\n(move to right\nto exclude)'), r, 0)
        self.order = ThumbListWidget(self) #QtGui.QListWidget()
      #  self.order.setDragDropMode(QtGui.QAbstractItemView.InternalMove)
        self.grid.addWidget(self.order, r, 1, 1, 2)
        self.ignore = ThumbListWidget(self) # QtGui.QListWidget()
      #  self.ignore.setDragDropMode(QtGui.QAbstractItemView.InternalMove)
        self.grid.addWidget(self.ignore, r, 3, 1, 2)
        r += 1
        self.log = QtGui.QLabel('')
        msg_palette = QtGui.QPalette()
        msg_palette.setColor(QtGui.QPalette.Foreground, QtCore.Qt.red)
        self.log.setPalette(msg_palette)
        self.grid.addWidget(self.log, r, 1, 1, 4)
        r += 1
        self.progressbar = QtGui.QProgressBar()
        self.progressbar.setMinimum(0)
        self.progressbar.setMaximum(10)
        self.progressbar.setValue(0)
        self.progressbar.setStyleSheet('QProgressBar {border: 1px solid grey; border-radius: 2px; text-align: center;}' \
                                       + 'QProgressBar::chunk { background-color: #6891c6;}')
        self.grid.addWidget(self.progressbar, r, 1, 1, 4)
        self.progressbar.setHidden(True)
        r += 1
        r += 1
        quit = QtGui.QPushButton('Done', self)
        self.grid.addWidget(quit, r, 0)
        quit.clicked.connect(self.quitClicked)
        QtGui.QShortcut(QtGui.QKeySequence('q'), self, self.quitClicked)
        pm = QtGui.QPushButton('Powermatch', self)
     #   pm.setMaximumWidth(wdth)
        self.grid.addWidget(pm, r, 1)
        pm.clicked.connect(self.pmClicked)
        pms = QtGui.QPushButton('Summary', self)
     #   pm.setMaximumWidth(wdth)
        self.grid.addWidget(pms, r, 2)
        pms.clicked.connect(self.pmClicked)
        opt = QtGui.QPushButton('Optimise', self)
        self.grid.addWidget(opt, r, 3)
        opt.clicked.connect(self.optClicked)
        help = QtGui.QPushButton('Help', self)
        help.setMaximumWidth(wdth)
        quit.setMaximumWidth(wdth)
        self.grid.addWidget(help, r, 5)
        help.clicked.connect(self.helpClicked)
        QtGui.QShortcut(QtGui.QKeySequence('F1'), self, self.helpClicked)
        try:
            ts = xlrd.open_workbook(self.scenarios + str(self.files[G].text()))
            ws = ts.sheet_by_name('Generators')
            self.setOrder(self.getGenerators(ws))
            ts.release_resources()
            del ts
        except:
            pass
        frame = QtGui.QFrame()
        frame.setLayout(self.grid)
        self.scroll = QtGui.QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setWidget(frame)
        self.layout = QtGui.QVBoxLayout(self)
        self.layout.addWidget(self.scroll)
    #    tab1.setLayout(self.layout)
    #    self.tabs.addTab(tab1,'Parms')
    #    self.tabs.addTab(tab2,'Constraints')
    #    self.tabs.addTab(tab3,'Generators')
    #    self.tabs.addTab(tab4,'Summary')
    #    self.tabs.addTab(tab5,'Details')
        self.setWindowTitle('SIREN - powermatch (' + fileVersion() + ') - Powermatch')
        self.setWindowIcon(QtGui.QIcon('sen_icon32.ico'))
        self.center()
        self.resize(int(self.sizeHint().width()* 1.07), int(self.sizeHint().height() * 1.1))
        self.show()

    def center(self):
        frameGm = self.frameGeometry()
        screen = QtGui.QApplication.desktop().screenNumber(QtGui.QApplication.desktop().cursor().pos())
        centerPoint = QtGui.QApplication.desktop().availableGeometry(screen).center()
        frameGm.moveCenter(centerPoint)
        self.move(frameGm.topLeft())

    def fileChanged(self):
        self.log.setText('')
        for i in range(5):
            if self.files[i].hasFocus():
                break
        curfile = self.scenarios + self.files[i].text()
        if i == R:
            if self.files[i].text() == '':
                curfile = self.scenarios + self.files[D].text()
                curfile = curfile.replace('data', 'results')
                curfile = curfile.replace('Data', 'Results')
            newfile = str(QtGui.QFileDialog.getSaveFileName(None, 'Save ' + self.file_labels[i] + ' file',
                      curfile, 'Excel Files (*.xlsx)'))
        else:
            newfile = str(QtGui.QFileDialog.getOpenFileName(self, 'Open ' + self.file_labels[i] + ' file',
                      curfile))
        if newfile != '':
            if i < D:
                ts = xlrd.open_workbook(newfile, on_demand = True)
                ndx = 0
                self.sheets[i].clear()
                j = -1
                for sht in ts.sheet_names():
                    j += 1
                    self.sheets[i].addItem(str(sht))
                    if str(sht) == self.file_labels[i]:
                        ndx = j
                self.sheets[i].setCurrentIndex(ndx)
                if i == 1:
                    ws = ts.sheet_by_index(ndx)
                    self.setOrder(self.getGenerators(ws))
                ts.release_resources()
                del ts
            if newfile[: len(self.scenarios)] == self.scenarios:
                self.files[i].setText(newfile[len(self.scenarios):])
            else:
                self.files[i].setText(newfile)
            if i == D and self.change_res:
                newfile = str(self.files[D].text())
                newfile = newfile.replace('data', 'results')
                newfile = newfile.replace('Data', 'Results')
                if newfile != str(self.files[D].text()):
                    self.files[R].setText(newfile)
            self.updated = True

    def helpClicked(self):
        dialog = displayobject.AnObject(QtGui.QDialog(), self.help,
                 title='Help for powermatch (' + fileVersion() + ')', section='powermatch')
        dialog.exec_()

    def cpchanged(self):
        self.updated = True
        self.carbon_price = self.carbon.value()

    def changes(self):
        self.updated = True

    def quitClicked(self):
        if self.updated:
            updates = {}
            lines = []
            lines.append('adjust_generators=' + str(self.adjust.isChecked()))
            lines.append('carbon_price=' + str(self.carbon_price))
            for i in range(len(self.file_labels)):
                lines.append(str(self.file_labels[i].lower()) + '_file=' + str(self.files[i].text()))
            for i in range(D):
                lines.append(str(self.file_labels[i].lower()) + '_sheet=' + str(self.sheets[i].currentText()))
            lines.append('optimise_population=' + str(self.optimise_population))
            lines.append('optimise_generations=' + str(self.optimise_generations))
            lines.append('optimise_stop=' + str(self.optimise_stop))
            updates['Powermatch'] = lines
            SaveIni(updates)
        self.close()

    def editClicked(self):
        self.log.setText('')
        it = self.file_labels.index(self.sender().text())
        try:
            if str(self.files[it].text()).find('/') >= 0:
                ts = xlrd.open_workbook(str(self.files[it].text()))
            else:
                ts = xlrd.open_workbook(self.scenarios + str(self.files[it].text()))
            try:
                sht = self.sheets[it].currentText()
            except:
                self.log.setText(self.sheets[it].currentText() + ' not found in ' + self.file_labels[it] + ' spreadsheet.')
                return
            ws = ts.sheet_by_name(sht)
        except:
            ts = None
            ws = None
        if it == C: # constraints
            try:
                constraints = self.getConstraints(ws)
            except:
                return
            sp_pts = [2] * 11
            sp_pts[4] = 3 # discharge loss
            sp_pts[6] = 3 # parasitic loss
            sp_pts[9] = 3 # recharge loss
            dialog = displaytable.Table(constraints, title=self.sender().text(),
                 save_folder=self.scenarios, edit=True, decpts=sp_pts, abbr=False)
            dialog.exec_()
            if dialog.getValues() is not None:
                constraints = dialog.getValues()
        elif it == G: # generators
            try:
                generators = self.getGenerators(ws)
            except:
                return
            sp_pts = [2] * 8
            sp_pts[7] = 0 # dispatch order column
            dialog = displaytable.Table(generators, title=self.sender().text(),
                 save_folder=self.scenarios, edit=True, decpts=sp_pts, abbr=False)
            dialog.exec_()
            if dialog.getValues() is not None:
                newgenerators = dialog.getValues()
                self.setOrder(generators)
        elif it == O: # optimisation
            try:
                optimisation = self.getoptimisation(ws, None)
            except:
                return
            dialog = displaytable.Table(optimisation, title=self.sender().text(),
                     save_folder=self.scenarios, edit=True)
            dialog.exec_()
            if dialog.getValues() is not None:
                optimisation = dialog.getValues()
        if ts is not None:
            ts.release_resources()
            del ts
        newfile = dialog.savedfile
        if newfile is not None:
            if newfile[: len(self.scenarios)] == self.scenarios:
                self.files[it].setText(newfile[len(self.scenarios):])
            else:
                self.files[it].setText(newfile)
            self.log.setText(self.file_labels[it] + ' spreadsheet changed.')

    def getConstraints(self, ws):
        if ws is None:
            constraints = {}
            constraints['<name>'] = Constraint('<name>', '<category>', 0., 1.,
                                              1., 1., 1., 0., 1., 0., 0.)
            return constraints
        if ws.cell_value(1, 0) == 'Name' and ws.cell_value(1, 1) == 'Category':
            cat_col = 1
            for col in range(ws.ncols):
                if ws.cell_value(0, col)[:8] == 'Capacity':
                    cap_col = [col, col + 1]
                elif ws.cell_value(0, col)[:9] == 'Ramp Rate':
                    ramp_col = [col, col + 1]
                elif ws.cell_value(0, col)[:8] == 'Recharge':
                    rec_col = [col, col + 1]
                elif ws.cell_value(0, col)[:9] == 'Discharge':
                    dis_col = [col, col + 1]
                elif ws.cell_value(1, col)[:9] == 'Parasitic':
                    par_col = col
            strt_row = 2
        elif ws.cell_value(0, 0) == 'Name': # saved file
            cap_col = [-1, -1]
            ramp_col = [-1, -1]
            rec_col = [-1, -1]
            dis_col = [-1, -1]
            for col in range(ws.ncols):
                if ws.cell_value(0, col)[:8] == 'Category':
                    cat_col = col
                elif ws.cell_value(0, col)[:8] == 'Capacity':
                    if ws.cell_value(0, col)[-3:] == 'Min':
                        cap_col[0] = col
                    else:
                        cap_col[1] = col
                elif ws.cell_value(0, col)[:6] == 'Rampup':
                    ramp_col[0] = col
                elif ws.cell_value(0, col)[:8] == 'Rampdown':
                    ramp_col[1] = col
                elif ws.cell_value(0, col)[:8] == 'Recharge':
                    if ws.cell_value(0, col)[-3:] == 'Max':
                        rec_col[0] = col
                    else:
                        rec_col[1] = col
                elif ws.cell_value(0, col)[:9] == 'Discharge':
                    if ws.cell_value(0, col)[-3:] == 'Max':
                        dis_col[0] = col
                    else:
                        dis_col[1] = col
                elif ws.cell_value(0, col)[:9] == 'Parasitic':
                    par_col = col
            strt_row = 1
        else:
            self.log.setText('Not a ' + self.file_labels[it] + ' worksheet.')
            return
        try:
            cat_col = cat_col
        except:
            self.log.setText('Not a ' + self.file_labels[it] + ' worksheet.')
            return
        constraints = {}
        for row in range(strt_row, ws.nrows):
            constraints[str(ws.cell_value(row, 0))] = Constraint(str(ws.cell_value(row, 0)), str(ws.cell_value(row, cat_col)),
                                     ws.cell_value(row, cap_col[0]), ws.cell_value(row, cap_col[1]),
                                     ws.cell_value(row, ramp_col[0]), ws.cell_value(row, ramp_col[1]),
                                     ws.cell_value(row, rec_col[0]), ws.cell_value(row, rec_col[1]),
                                     ws.cell_value(row, dis_col[0]), ws.cell_value(row, dis_col[1]),
                                     ws.cell_value(row, par_col))
        return constraints

    def getGenerators(self, ws):
        if ws is None:
            generators = {}
            generators['<name>'] = Facility('<name>', 0, '<constraint>', 0., 0., 0., 0.)
            return generators
        if ws.cell_value(0, 0) != 'Name':
            self.log.setText('Not a ' + self.file_labels[G] + ' worksheet.')
            return
        for col in range(ws.ncols):
            if ws.cell_value(0, col)[:8] == 'Dispatch' or ws.cell_value(0, col) == 'Order':
                ord_col = col
            elif ws.cell_value(0, col) == 'Constraint':
                con_col = col
            elif ws.cell_value(0, col)[:8] == 'Capacity':
                cap_col = col
            elif ws.cell_value(0, col)[:7] == 'Initial':
                ini_col = col
            elif ws.cell_value(0, col) == 'LCOE CF':
                lcc_col = col
            elif ws.cell_value(0, col)[:4] == 'LCOE':
                lco_col = col
            elif ws.cell_value(0, col) == 'Emissions':
                emi_col = col
        try:
            lco_col = lco_col
        except:
            self.log.setText('Not a ' + self.file_labels[G] + ' worksheet.')
            return
        generators = {}
        for row in range(1, ws.nrows):
            generators[str(ws.cell_value(row, 0))] = Facility(str(ws.cell_value(row, 0)),
                                     ws.cell_value(row, ord_col), str(ws.cell_value(row, con_col)),
                                     ws.cell_value(row, cap_col), ws.cell_value(row, lco_col),
                                     ws.cell_value(row, lcc_col), ws.cell_value(row, emi_col),
                                     initial=ws.cell_value(row, ini_col))
        return generators

    def getoptimisation(self, ws, generators):
        if ws is None:
            optimisation = {}
            optimisation['<name>'] = Optimisation('<name>', 'None', None)
            return optimisation
        if ws.cell_value(0, 0) != 'Name':
            self.log.setText('Not an ' + self.file_labels[O] + ' worksheet.')
            return
        cols = ['Name', 'Approach', 'Values', 'Capacity Max', 'Capacity Min',
                'Capacity Step', 'Capacities']
        coln = [-1] * len(cols)
        for col in range(ws.ncols):
            try:
                i = cols.index(ws.cell_value(0, col))
                coln[i] = col
            except:
                pass
        if coln[0] < 0:
            self.log.setText('Not an ' + self.file_labels[O] + ' worksheet.')
            return
        optimisation = {}
        for row in range(1, ws.nrows):
            tech = str(ws.cell_value(row, 0))
            if coln[2] > 0: # values format
                optimisation[tech] = Optimisation(tech,
                                     ws.cell_value(row, coln[1]),
                                     ws.cell_value(row, coln[2]))
            else:
                if ws.cell_value(row, coln[1]) == 'Discrete': # fudge values format
                    optimisation[tech] = Optimisation(tech,
                                         ws.cell_value(row, coln[1]),
                                         ws.cell_value(row, coln[-1]))
                else:
                    optimisation[tech] = Optimisation(tech, '', '')
                    for col in range(1, len(coln)):
                        if coln[col] > 0:
                            attr = cols[col].lower().replace(' ', '_')
                            setattr(optimisation[tech], attr,
                                    ws.cell_value(row, coln[col]))
            try:
                optimisation[tech].capacity = generators[tech].capacity
            except:
                pass
        return optimisation

    def setOrder(self, generators=None):
        self.order.clear()
        self.ignore.clear()
        if generators is None:
            order = ['Storage', 'Biomass', 'PHS', 'Gas', 'CCG1', 'Other', 'Coal']
            for stn in order:
                self.order.addItem(stn)
        else:
            order = []
            zero = []
            for key, value in generators.items():
                if value.capacity == 0:
                    continue
                try:
                    o = int(value.order)
                    if o > 0:
                        while len(order) <= o:
                            order.append([])
                        order[o - 1].append(key)
                    elif o == 0:
                        zero.append(key)
                except:
                    pass
            order.append(zero)
            for cat in order:
                for stn in cat:
                    self.order.addItem(stn)

    def pmClicked(self):
        self.log.setText(self.sender().text() + ' processing started')
        if self.sender().text() == 'Summary': # summary only?
            summ_only = True
        else:
            summ_only = False
        self.progressbar.setMinimum(0)
        self.progressbar.setMaximum(10)
        self.progressbar.setHidden(False)
        details = self.details
        try:
            if str(self.files[C].text()).find('/') >= 0:
                ts = xlrd.open_workbook(str(self.files[C].text()))
            else:
                ts = xlrd.open_workbook(self.scenarios + str(self.files[C].text()))
            ws = ts.sheet_by_name(self.sheets[C].currentText())
            constraints = self.getConstraints(ws)
            ts.release_resources()
            del ts
        except:
            self.log.setText('Error accessing Constraints.')
            self.progressbar.setHidden(True)
            return
        try:
            if str(self.files[G].text()).find('/') >= 0:
                ts = xlrd.open_workbook(str(self.files[G].text()))
            else:
                ts = xlrd.open_workbook(self.scenarios + str(self.files[G].text()))
            ws = ts.sheet_by_name(self.sheets[G].currentText())
            generators = self.getGenerators(ws)
            ts.release_resources()
            del ts
        except:
            self.log.setText('Error accessing Generators.')
            self.progressbar.setHidden(True)
            return
        start_time = time.time()
        re_capacities = [0.] * len(tech_names)
        if str(self.files[D].text()).find('/') >= 0:
            pm_data_file = str(self.files[D].text())
        else:
            pm_data_file = self.scenarios + str(self.files[D].text())
        re_capacity = 0.
        data = []
        load = []
        shortfall = []
        if pm_data_file[-4:] == '.xls': #xls format
            xlsx = False
            details = False
            ts = xlrd.open_workbook(pm_data_file)
            ws = ts.sheet_by_index(0)
            if ws.cell_value(0, 0) != 'Hourly Shortfall Table' \
              or ws.cell_value(0, 4) != 'Generation Summary Table':
                self.log.setText('not a Powerbalance spreadsheet')
                self.progressbar.setHidden(True)
                return
            for row in range(20):
                if ws.cell_value(row, 5) == 'Total':
                    re_capacity = ws.cell_value(row, 6)
                    re_generation = ws.cell_value(row, 8)
                    break
            for row in range(2, ws.nrows):
                shortfall.append(ws.cell_value(row, 2))
        else: # xlsx format
            xlsx = True
            shortfall = [0.] * 8760
            if details:
                cols = []
            ts = oxl.load_workbook(pm_data_file)
            ws = ts.active
            top_row = ws.max_row - 8760
            if ws.cell(row=top_row, column=1).value != 'Hour' or ws.cell(row=top_row, column=2).value != 'Period':
                self.log.setText('not a Powermatch data spreadsheet')
                self.progressbar.setHidden(True)
                return
            typ_row = top_row - 1
            while typ_row > 0:
                if ws.cell(row=typ_row, column=3).value in tech_names:
                    break
                typ_row -= 1
            else:
                self.log.setText('no suitable data')
                return
            icap_row = typ_row + 1
            while icap_row < top_row:
                if ws.cell(row=icap_row, column=1).value == 'Capacity (MW)':
                    break
                icap_row += 1
            else:
                self.log.setText('no capacity data')
                return
       #     adjustby = None
            if self.adjust.isChecked():
                adjustin = []
                for col in range(3, ws.max_column + 1):
                    try:
                        valu = ws.cell(row=typ_row, column=col).value.replace('-','')
                        i = tech_names.index(valu)
                    except:
                        break
                    if valu == 'Load':
                        adjustin.append([tech_names[i], 0])
                    else:
                        try:
                            adjustin.append([tech_names[i], float(ws.cell(row=icap_row, column=col).value)])
                        except:
                            pass
                for i in range(self.order.count()):
                    if generators[self.order.item(i).text()].capacity > 0:
                        adjustin.append([self.order.item(i).text(),
                                        generators[self.order.item(i).text()].capacity])
                adjust = Adjustments(adjustin, self.adjustby)
                adjust.exec_()
                self.adjustby = adjust.getValues()
                if self.adjustby is None:
                    self.log.setText('Execution aborted.')
                    self.progressbar.setHidden(True)
                    return
            load_col = -1
            det_col = 3
            self.progressbar.setValue(1)
            for col in range(3, ws.max_column + 1):
                try:
                    valu = ws.cell(row=typ_row, column=col).value.replace('-','')
                    i = tech_names.index(valu)
                except:
                    break # ?? or continue??
                if tech_names[i] != 'Load':
                    try:
                        if ws.cell(row=icap_row, column=col).value <= 0:
                            continue
                    except:
                        continue
                data.append([])
                try:
                    multiplier = self.adjustby[tech_names[i]]
                except:
                    multiplier = 1.
                if details:
                    cols.append(i)
                    try:
                        re_capacities[i] = ws.cell(row=icap_row, column=col).value * multiplier
                    except:
                        pass
                try:
                    re_capacity += ws.cell(row=icap_row, column=col).value * multiplier
                except:
                    pass
                for row in range(top_row + 1, ws.max_row + 1):
                    data[-1].append(ws.cell(row=row, column=col).value * multiplier)
                    try:
                        if tech_names[i] == 'Load':
                            load_col = len(data) - 1
                            shortfall[row - top_row - 1] += ws.cell(row=row, column=col).value * multiplier
                        else:
                            shortfall[row - top_row - 1] -= ws.cell(row=row, column=col).value * multiplier
                    except:
                        pass
            ts.close()
        if str(self.files[R].text()) == '':
            i = pm_data_file.find('/')
            if i >= 0:
                data_file = pm_data_file[i + 1:]
            else:
                data_file = pm_data_file
            data_file = data_file.replace('data', 'results')
        else:
            if str(self.files[R].text()).find('/') >= 0:
                data_file = str(self.files[R].text())
            else:
                data_file = self.scenarios + str(self.files[R].text())
        if not xlsx:
            data_file += 'x'
        self.progressbar.setValue(2)
        headers = ['Facility', 'Capacity (MW)', 'Subtotal (MWh)', 'CF', 'Cost ($)', 'LCOE ($/MWh)', 'Emissions (tCO2e)']
        sp_cols = []
        sp_cap = []
        if summ_only: # summary only?
            sp_data = []
            sp_gen = []
            sp_load = 0.
            for i in cols:
                if tech_names[i] == 'Load':
                    sp_load = sum(data[load_col])
                    continue
                sp_data.append([tech_names[i], re_capacities[i], 0., '', '', '', ''])
                for g in data[len(sp_data)]:
                    sp_data[-1][2] += g
                # if ignore not used
        else: # normal
            ds = oxl.Workbook()
            ns = ds.active
            ns.title = 'Detail'
            ss = ds.create_sheet('Summary', 0)
            re_sum = '=('
            cap_row = 1
            ns.cell(row=cap_row, column=2).value = headers[1]
            ss.cell(row=3, column=1).value = headers[0]
            ss.cell(row=3, column=2).value = headers[1]
            ini_row = 2
            ns.cell(row=ini_row, column=2).value = 'Initial Capacity'
            sum_row = 3
            ns.cell(row=sum_row, column=2).value = headers[2]
            ss.cell(row=3, column=3).value = headers[2]
            cf_row = 4
            ns.cell(row=cf_row, column=2).value = headers[3]
            ss.cell(row=3, column=4).value = headers[3]
            cost_row = 5
            ns.cell(row=cost_row, column=2).value = headers[4]
            ss.cell(row=3, column=5).value = headers[4]
            lcoe_row = 6
            ns.cell(row=lcoe_row, column=2).value = headers[5]
            ss.cell(row=3, column=6).value = headers[5]
            emi_row = 7
            ns.cell(row=emi_row, column=2).value = headers[6]
            ss.cell(row=3, column=7).value = headers[6]
            ss_row = 3
            fall_row = 8
            ns.cell(row=fall_row, column=2).value = 'Shortfall periods'
            what_row = 9
            hrows = 10
            ns.cell(row=what_row, column=1).value = 'Hour'
            ns.cell(row=what_row, column=2).value = 'Period'
            ns.cell(row=what_row, column=3).value = 'Load'
            o = 4
            col = 3
            if details:
                for i in cols:
                    if tech_names[i] == 'Load':
                        continue
                    col += 1
                    sp_cols.append(tech_names[i])
                    sp_cap.append(re_capacities[i])
                    ss_row += 1
                    ns.cell(row=what_row, column=col).value = tech_names[i]
                    ss.cell(row=ss_row, column=1).value = '=Detail!' + ss_col(col) + str(what_row)
                    ns.cell(row=cap_row, column=col).value = re_capacities[i]
                    ns.cell(row=cap_row, column=col).number_format = '#,##0.00'
                    ss.cell(row=ss_row, column=2).value = '=Detail!' + ss_col(col) + str(cap_row)
                    ss.cell(row=ss_row, column=2).number_format = '#,##0.00'
                    ns.cell(row=sum_row, column=col).value = '=SUM(' + ss_col(col) \
                            + str(hrows) + ':' + ss_col(col) + str(hrows + 8759) + ')'
                    ns.cell(row=sum_row, column=col).number_format = '#,##0'
                    ss.cell(row=ss_row, column=3).value = '=Detail!' + ss_col(col) + str(sum_row)
                    ss.cell(row=ss_row, column=3).number_format = '#,##0'
                    re_sum += 'C' + str(ss_row) + '+'
                    ns.cell(row=cf_row, column=col).value = '=' + ss_col(col) + '3/' \
                            + ss_col(col) + '1/8760'
                    ns.cell(row=cf_row, column=col).number_format = '#,##0.00'
                    ss.cell(row=ss_row, column=4).value = '=Detail!' + ss_col(col) + str(cf_row)
                    ss.cell(row=ss_row, column=4).number_format = '#,##0.00'
                    if tech_names[i] not in generators:
                        continue
                    if generators[tech_names[i]].lcoe > 0:
                        ns.cell(row=cost_row, column=col).value = generators[tech_names[i]].lcoe * \
                                generators[tech_names[i]].lcoe_cf * 8760 * re_capacities[i]
                        ns.cell(row=cost_row, column=col).number_format = '$#,##0'
                        ss.cell(row=ss_row, column=5).value = '=Detail!' + ss_col(col) + str(cost_row)
                        ss.cell(row=ss_row, column=5).number_format = '$#,##0'
                        ns.cell(row=lcoe_row, column=col).value = '=IF(AND(' + ss_col(col) + str(cf_row) + '>0,' \
                                + ss_col(col) + str(cap_row) + '>0),' + ss_col(col) + str(cost_row) + '/8760/' \
                                + ss_col(col) + str(cf_row) +'/' + ss_col(col) + str(cap_row) + ',"")'
                        ns.cell(row=lcoe_row, column=col).number_format = '$#,##0.00'
                        ss.cell(row=ss_row, column=6).value = '=Detail!' + ss_col(col) + str(lcoe_row)
                        ss.cell(row=ss_row, column=6).number_format = '$#,##0.00'
                    if generators[tech_names[i]].emissions > 0:
                        ns.cell(row=emi_row, column=col).value = '=' + ss_col(col) + str(sum_row) \
                                + '*' + str(generators[tech_names[i]].emissions)
                        ns.cell(row=emi_row, column=col).number_format = '#,##0'
                        ss.cell(row=ss_row, column=7).value = '=Detail!' + ss_col(col) + str(emi_row)
                        ss.cell(row=ss_row, column=7).number_format = '#,##0'
                shrt_col = col + 1
            else:
                shrt_col = 5
                sp_cols.append('Renewable')
                sp_cap.append(re_capacity)
                ns.cell(row=what_row, column=4).value = 'Renewable'
                ns.cell(row=cap_row, column=4).value = re_capacity
                ns.cell(row=cap_row, column=4).number_format = '#,##0.00'
                if xlsx:
                    ns.cell(row=sum_row, column=4).value = '=SUM(D' + str(hrows) + ':D' + str(hrows + 8759) + ')'
                else:
                    ns.cell(row=sum_row, column=4).value = re_generation
                ns.cell(row=sum_row, column=4).number_format = '#,##0'
                ns.cell(row=cf_row, column=4).value = '=D3/D1/8760'
                ns.cell(row=cf_row, column=4).number_format = '#,##0.00'
                ss_row += 1
                ss.cell(row=ss_row, column=1).value = '=Detail!D' + str(what_row)
                ss.cell(row=ss_row, column=2).value = '=Detail!D' + str(cap_row)
                ss.cell(row=ss_row, column=2).number_format = '#,##0.00'
                ss.cell(row=ss_row, column=3).value = '=Detail!D' + str(sum_row)
                ss.cell(row=ss_row, column=3).number_format = '#,##0'
                ss.cell(row=ss_row, column=4).value = '=Detail!D' + str(cf_row)
                ss.cell(row=ss_row, column=4).number_format = '#,##0.00'
            ns.cell(row=fall_row, column=shrt_col).value = '=COUNTIF(' + ss_col(shrt_col) \
                            + str(hrows) + ':' + ss_col(shrt_col) + str(hrows + 8759) + ',">0")'
            ns.cell(row=fall_row, column=shrt_col).number_format = '#,##0'
            ns.cell(row=what_row, column=shrt_col).value = 'Shortfall'
            for col in range(3, shrt_col + 1):
                ns.cell(row=what_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                        vertical='bottom', horizontal='center')
            for row in range(hrows, 8760 + hrows):
                if xlsx:
                    ns.cell(row=row, column=1).value = ws.cell(row=top_row + row - hrows + 1, column=1).value
                    ns.cell(row=row, column=2).value = ws.cell(row=top_row + row - hrows + 1, column=2).value
                    ns.cell(row=row, column=3).value = round(data[load_col][row - hrows], 2)
                    if not details:
                        ns.cell(row=row, column=4).value = round(data[load_col][row - hrows] - shortfall[row - hrows], 2)
                else:
                    ns.cell(row=row, column=1).value = ws.cell_value(row - hrows + 2, 0)
                    ns.cell(row=row, column=2).value = ws.cell_value(row - hrows + 2, 1)
                ns.cell(row=row, column=shrt_col).value = round(shortfall[row - hrows], 2)
                for col in range(3, shrt_col + 1):
                    ns.cell(row=row, column=col).number_format = '#,##0.00'
            if details:
                col = 3
                for i in range(len(data)):
                    if cols[i] == load_col:
                        continue
                    col += 1
                    for row in range(hrows, 8760 + hrows):
                        ns.cell(row=row, column=col).value = data[i][row - hrows]
            col = shrt_col + 1
        order = [] #'Storage', 'Biomass', 'PHS', 'Gas', 'CCG1', 'Other', 'Coal']
        for itm in range(self.order.count()):
            order.append(str(self.order.item(itm).text()))
        #storage? = []
        self.progressbar.setValue(3)
        for gen in order:
            try:
                capacity = generators[gen].capacity * self.adjustby[gen]
            except:
                capacity = generators[gen].capacity
            if generators[gen].constraint in constraints and \
              constraints[generators[gen].constraint].category == 'Storage': # storage
                storage = [0., 0., 0., 0.] # capacity, initial, min level, max drain
                storage[0] = capacity
                if not summ_only:
                    ns.cell(row=cap_row, column=col).value = round(capacity, 2)
                    ns.cell(row=cap_row, column=col).number_format = '#,##0.00'
                storage[1] = generators[gen].initial
                if constraints[generators[gen].constraint].capacity_min > 0:
                    storage[2] = capacity * constraints[generators[gen].constraint].capacity_min
                if constraints[generators[gen].constraint].capacity_max > 0:
                    storage[3] = capacity * constraints[generators[gen].constraint].capacity_max
                else:
                    storage[3] = capacity
                recharge = [0., 0.] # cap, loss
                if constraints[generators[gen].constraint].recharge_max > 0:
                    recharge[0] = capacity * constraints[generators[gen].constraint].recharge_max
                else:
                    recharge[0] = capacity
                if constraints[generators[gen].constraint].recharge_loss > 0:
                    recharge[1] = constraints[generators[gen].constraint].recharge_loss
                discharge = [0., 0.] # cap, loss
                if constraints[generators[gen].constraint].discharge_max > 0:
                    discharge[0] = capacity * constraints[generators[gen].constraint].discharge_max
                if constraints[generators[gen].constraint].discharge_loss > 0:
                    discharge[1] = constraints[generators[gen].constraint].discharge_loss
                if constraints[generators[gen].constraint].parasitic_loss > 0:
                    parasite = constraints[generators[gen].constraint].parasitic_loss / 24.
                else:
                    parasite = 0.
                storage_carry = storage[1] # generators[gen].initial
                if not summ_only:
                    ns.cell(row=ini_row, column=col).value = round(storage_carry, 2)
                    ns.cell(row=ini_row, column=col).number_format = '#,##0.00'
                storage_bal = []
                storage_can = 0.
                for row in range(8760):
                    if storage_carry > 0:
                        storage_carry = storage_carry * (1 - parasite)
                    storage_loss = 0.
                    if shortfall[row] < 0:  # excess generation
                        can_use = - (storage[0] - storage_carry) * (1 / (1 - recharge[1]))
                        if can_use < 0: # can use some
                            if shortfall[row] > can_use:
                                can_use = shortfall[row]
                            if can_use < - recharge[0] * (1 / (1 - recharge[1])):
                                can_use = - recharge[0]
                        else:
                            can_use = 0.
                        storage_carry -= (can_use * (1 - recharge[1]))
                        shortfall[row] -= can_use
                    else: # shortfall
                        can_use = shortfall[row] * (1 / (1 - discharge[1]))
                        if can_use > storage_carry - storage[2]:
                            can_use = storage_carry - storage[2]
                        if can_use > 0:
                            storage_loss = can_use * discharge[1]
                            storage_carry -= can_use
                            shortfall[row] -= (can_use - storage_loss)
                            if storage_carry < 0:
                                storage_carry = 0
                        else:
                            can_use = 0.
                    storage_bal.append(storage_carry)
                    if not summ_only:
                        ns.cell(row=row + hrows, column=col).value = round(can_use, 2)
                        ns.cell(row=row + hrows, column=col + 1).value = round(storage_carry, 2)
                        ns.cell(row=row + hrows, column=col + 2).value = round(shortfall[row], 2)
                        ns.cell(row=row + hrows, column=col).number_format = '#,##0.00'
                        ns.cell(row=row + hrows, column=col + 1).number_format = '#,##0.00'
                        ns.cell(row=row + hrows, column=col + 2).number_format = '#,##0.00'
                    else:
                        if can_use > 0:
                            storage_can += can_use
                if not summ_only:
                    ns.cell(row=sum_row, column=col).value = '=SUMIF(' + ss_col(col) + \
                            str(hrows) + ':' + ss_col(col) + str(hrows + 8759) + ',">0")'
                    ns.cell(row=sum_row, column=col).number_format = '#,##0'
                    ns.cell(row=cf_row, column=col).value = '=' + ss_col(col) + '3/' + ss_col(col) + '1/8760'
                    ns.cell(row=cf_row, column=col).number_format = '#,##0.00'
                    col += 3
                else:
                    sp_data.append([gen, storage[0], storage_can, '', '', '', ''])
            else:
                if not summ_only:
                    ns.cell(row=cap_row, column=col).value = round(capacity, 2)
                    ns.cell(row=cap_row, column=col).number_format = '#,##0.00'
                    for row in range(8760):
                        if shortfall[row] >= 0: # shortfall?
                            if shortfall[row] >= capacity:
                                shortfall[row] = shortfall[row] - capacity
                                ns.cell(row=row + hrows, column=col).value = round(capacity, 2)
                            else:
                                ns.cell(row=row + hrows, column=col).value = round(shortfall[row], 2)
                                shortfall[row] = 0
                        else:
                            ns.cell(row=row + hrows, column=col).value = 0
                        ns.cell(row=row + hrows, column=col + 1).value = round(shortfall[row], 2)
                        ns.cell(row=row + hrows, column=col).number_format = '#,##0.00'
                        ns.cell(row=row + hrows, column=col + 1).number_format = '#,##0.00'
                    ns.cell(row=sum_row, column=col).value = '=SUM(' + ss_col(col) + str(hrows) + \
                            ':' + ss_col(col) + str(hrows + 8759) + ')'
                    ns.cell(row=sum_row, column=col).number_format = '#,##0'
                    ns.cell(row=cf_row, column=col).value = '=' + ss_col(col) + '3/' + ss_col(col) + '1/8760'
                    ns.cell(row=cf_row, column=col).number_format = '#,##0.00'
                    col += 2
                else:
                    gen_can = 0.
                    for row in range(8760):
                        if shortfall[row] >= 0: # shortfall?
                            if shortfall[row] >= capacity:
                                shortfall[row] = shortfall[row] - capacity
                                gen_can += capacity
                            else:
                                gen_can += shortfall[row]
                                shortfall[row] = 0
                    sp_data.append([gen, capacity, gen_can, '', '', '', ''])
        self.progressbar.setValue(4)
        if summ_only:
            cap_sum = 0.
            gen_sum = 0.
            re_sum = 0.
            cost_sum = 0.
            co2_sum = 0.
            for sp in range(len(sp_data)):
                if sp_data[sp][1] > 0:
                    cap_sum += sp_data[sp][1]
                    sp_data[sp][3] = sp_data[sp][2] / sp_data[sp][1] / 8760
                gen_sum += sp_data[sp][2]
                gen = sp_data[sp][0]
                if gen in tech_names:
                    re_sum += sp_data[sp][2]
                if gen not in generators:
                    continue
                if generators[gen].lcoe > 0:
                    sp_data[sp][4] = generators[gen].lcoe * generators[gen].lcoe_cf * 8760 * sp_data[sp][1]
                    if sp_data[sp][1] > 0 and sp_data[sp][3] > 0:
                        sp_data[sp][5] = sp_data[sp][4] / 8760 / sp_data[sp][3] / sp_data[sp][1]
                    cost_sum += sp_data[sp][4]
                if generators[gen].emissions > 0:
                    sp_data[sp][6] = sp_data[sp][2] * generators[gen].emissions
                    co2_sum += sp_data[sp][6]
            if cap_sum > 0:
                cs = gen_sum / cap_sum / 8760
            else:
                cs = ''
            if gen_sum > 0:
                gs = cost_sum / gen_sum
            else:
                gs = ''
            sp_data.append(['Total', cap_sum, gen_sum, cs, cost_sum, gs, co2_sum])
            if self.carbon_price > 0:
                cc = co2_sum * self.carbon_price
                cs = (cost_sum + cc) / gen_sum
                if self.carbon_price == int(self.carbon_price):
                   cp = str(int(self.carbon_price))
                else:
                   cp = '{:.2f}'.format(self.carbon_price)
                sp_data.append(['Carbon Cost ($' + cp + ')', '', '', '', cc, cs])
            sp_data.append(['RE %age', round(re_sum * 100. / gen_sum, 1)])
            sf_sums = [0., 0., 0.]
            for sf in range(len(shortfall)):
                if shortfall[sf] > 0:
                    sf_sums[0] += shortfall[sf]
                    sf_sums[2] += data[0][sf]
                else:
                    sf_sums[1] += shortfall[sf]
                    sf_sums[2] += data[0][sf]
            sp_data.append(' ')
            sp_data.append('Load Analysis')
            pct = '{:.1%})'.format((sf_sums[2] - sf_sums[0]) / sp_load)
            sp_data.append(['Load met (' + pct, '', sf_sums[2] - sf_sums[0],])
            pct = '{:.1%})'.format(sf_sums[0] / sp_load)
            sp_data.append(['Shortfall (' + pct, '', sf_sums[0]])
            sp_data.append(['Total Load', '', sp_load])
            pct = '{:.1%})'.format( -sf_sums[1] / sp_load)
            sp_data.append(['Surplus (' + pct, '', -sf_sums[1]])
            adjusted = False
            if self.adjustby is not None:
                for key, value in iter(sorted(self.adjustby.items())):
                    if value != 1:
                        if not adjusted:
                            adjusted = True
                            sp_data.append('Generators Adjustments:')
                        sp_data.append([key, value])
            list(map(list, list(zip(*sp_data))))
            sp_pts = [0, 2, 0, 2, 0, 2, 0]
            dialog = displaytable.Table(sp_data, title=self.sender().text(), fields=headers,
                     save_folder=self.scenarios, sortby='', decpts=sp_pts)
            dialog.exec_()
            self.progressbar.setValue(10)
            self.log.setText(self.sender().text() + ' completed')
            self.progressbar.setHidden(True)
            self.progressbar.setValue(0)
            return
        for column_cells in ns.columns:
            length = 0
            value = ''
            row = 0
            for cell in column_cells:
                if cell.row >= 0:
                    if len(str(cell.value)) > length:
                        length = len(str(cell.value))
                        value = cell.value
                        row = cell.row
            try:
                ns.column_dimensions[column_cells[0].column].width = max(length, 10)
            except:
                ns.column_dimensions[ss_col(column_cells[0].column)].width = max(length, 10)
        ns.column_dimensions['A'].width = 6
        col = shrt_col + 1
        for gen in order:
            ss_row += 1
            ns.cell(row=what_row, column=col).value = gen
            ss.cell(row=ss_row, column=1).value = '=Detail!' + ss_col(col) + str(what_row)
            ss.cell(row=ss_row, column=2).value = '=Detail!' + ss_col(col) + str(cap_row)
            ss.cell(row=ss_row, column=2).number_format = '#,##0.00'
            ss.cell(row=ss_row, column=3).value = '=Detail!' + ss_col(col) + str(sum_row)
            ss.cell(row=ss_row, column=3).number_format = '#,##0'
            ss.cell(row=ss_row, column=4).value = '=Detail!' + ss_col(col) + str(cf_row)
            ss.cell(row=ss_row, column=4).number_format = '#,##0.00'
            if generators[gen].lcoe > 0:
                try:
                    capacity = generators[gen].capacity * self.adjustby[gen]
                except:
                   capacity = generators[gen].capacity
                ns.cell(row=cost_row, column=col).value = generators[gen].lcoe * \
                        generators[gen].lcoe_cf * 8760 * capacity
                ns.cell(row=cost_row, column=col).number_format = '$#,##0'
                ss.cell(row=ss_row, column=5).value = '=Detail!' + ss_col(col) + str(cost_row)
                ss.cell(row=ss_row, column=5).number_format = '$#,##0'
                ns.cell(row=lcoe_row, column=col).value = '=IF(AND(' + ss_col(col) + str(cf_row) + '>0,' \
                            + ss_col(col) + str(cap_row) + '>0),' + ss_col(col) + str(cost_row) + '/8760/' \
                            + ss_col(col) + str(cf_row) + '/' + ss_col(col) + str(cap_row)+  ',"")'
                ns.cell(row=lcoe_row, column=col).number_format = '$#,##0.00'
                ss.cell(row=ss_row, column=6).value = '=Detail!' + ss_col(col) + str(lcoe_row)
                ss.cell(row=ss_row, column=6).number_format = '$#,##0.00'
            if generators[gen].emissions > 0:
                ns.cell(row=emi_row, column=col).value = '=' + ss_col(col) + str(sum_row) \
                        + '*' + str(generators[gen].emissions)
                ns.cell(row=emi_row, column=col).number_format = '#,##0'
                ss.cell(row=ss_row, column=7).value = '=Detail!' + ss_col(col) + str(emi_row)
                ss.cell(row=ss_row, column=7).number_format = '#,##0'
            ns.cell(row=what_row, column=col).alignment = oxl.styles.Alignment(wrap_text=True,
                    vertical='bottom', horizontal='center')
            ns.cell(row=what_row, column=col + 1).alignment = oxl.styles.Alignment(wrap_text=True,
                    vertical='bottom', horizontal='center')
            if constraints[generators[gen].constraint].category == 'Storage':
                ns.cell(row=what_row, column=col + 1).value = gen + '\nBalance'
                ns.cell(row=what_row, column=col + 2).value = 'After\n' + gen
                ns.cell(row=what_row, column=col + 2).alignment = oxl.styles.Alignment(wrap_text=True,
                        vertical='bottom', horizontal='center')
                ns.cell(row=fall_row, column=col + 2).value = '=COUNTIF(' + ss_col(col + 2) \
                        + str(hrows) + ':' + ss_col(col + 2) + str(hrows + 8759) + ',">0")'
                ns.cell(row=fall_row, column=col + 2).number_format = '#,##0'
                col += 3
            else:
                ns.cell(row=what_row, column=col + 1).value = 'After\n' + gen
                ns.cell(row=fall_row, column=col + 1).value = '=COUNTIF(' + ss_col(col + 1) \
                        + str(hrows) + ':' + ss_col(col + 1) + str(hrows + 8759) + ',">0")'
                ns.cell(row=fall_row, column=col + 1).number_format = '#,##0'
                col += 2
        self.progressbar.setValue(6)
        ns.row_dimensions[what_row].height = 30
        ns.freeze_panes = 'C' + str(hrows)
        ns.activeCell = 'C' + str(hrows)
        ss.cell(row=1, column=1).value = 'Powermatch - Summary'
        bold = oxl.styles.Font(bold=True, name=ss.cell(row=1, column=1).font.name)
        ss.cell(row=1, column=1).font = bold
        ss_row +=1
        for col in range(1, 8):
            ss.cell(row=3, column=col).font = bold
            ss.cell(row=ss_row, column=col).font = bold
        ss.cell(row=ss_row, column=2).value = '=SUM(B4:B' + str(ss_row - 1) + ')'
        ss.cell(row=ss_row, column=2).number_format = '#,##0.00'
        ss.cell(row=ss_row, column=3).value = '=SUM(C4:C' + str(ss_row - 1) + ')'
        ss.cell(row=ss_row, column=3).number_format = '#,##0'
        ss.cell(row=ss_row, column=4).value = '=C' + str(ss_row) + '/B' + str(ss_row) + '/8760'
        ss.cell(row=ss_row, column=4).number_format = '#,##0.00'
        ss.cell(row=ss_row, column=5).value = '=SUM(E4:E' + str(ss_row - 1) + ')'
        ss.cell(row=ss_row, column=5).number_format = '$#,##0'
        ss.cell(row=ss_row, column=6).value = '=E' + str(ss_row) + '/C' + str(ss_row)
        ss.cell(row=ss_row, column=6).number_format = '$#,##0.00'
        ss.cell(row=ss_row, column=7).value = '=SUM(G4:G' + str(ss_row - 1) + ')'
        ss.cell(row=ss_row, column=7).number_format = '#,##0'
        ss.cell(row=ss_row, column=1).value = 'Total'
        for column_cells in ss.columns:
            length = 0
            value = ''
            row = 0
            for cell in column_cells:
                if len(str(cell.value)) > length:
                    length = len(str(cell.value))
                    value = cell.value
                    row = cell.row
            try:
                ss.column_dimensions[column_cells[0].column].width = length * 1.15
            except:
                ss.column_dimensions[ss_col(column_cells[0].column)].width = length * 1.15
        ss.column_dimensions['D'].width = 7
        ss.column_dimensions['E'].width = 18
        last_col = ss_col(ns.max_column)
        r = 1
        if self.carbon_price > 0:
            r = 2
            ss_row += 1
            if self.carbon_price == int(self.carbon_price):
               cp = str(int(self.carbon_price))
            else:
               cp = '{:.2f}'.format(self.carbon_price)
            ss.cell(row=ss_row, column=1).value = 'Carbon Cost ($' + cp + ')'
            ss.cell(row=ss_row, column=5).value = '=G' + str(ss_row - 1) + '*' + \
                    str(self.carbon_price)
            ss.cell(row=ss_row, column= 5).number_format = '$#,##0'
            ss.cell(row=ss_row, column=6).value = '=(E' + str(ss_row - 1) + \
                    '+E'  + str(ss_row) + ')/C' + str(ss_row - 1)
            ss.cell(row=ss_row, column=6).number_format = '$#,##0.00'
        ss_row += 1
        ss.cell(row=ss_row, column=1).value = 'RE %age'
        ss.cell(row=ss_row, column=2).value = re_sum[:-1] + ')/C' + str(ss_row - r)
        ss.cell(row=ss_row, column=2).number_format = '#,##0.0%'
        ss_row += 2
        ss.cell(row=ss_row, column=1).value = 'Load Analysis'
        ss.cell(row=ss_row, column=1).font = bold
        ss_row += 1
        ss.cell(row=ss_row, column=1).value = 'Load met'
        ss.cell(row=ss_row, column=3).value = '=SUMIF(Detail!' + last_col + str(hrows) + ':Detail!' \
            + last_col + str(hrows + 8759) + ',"<=0",Detail!C' + str(hrows) + ':Detail!C' \
            + str(hrows + 8759) + ')+SUMIF(Detail!' + last_col + str(hrows) + ':Detail!' \
            + last_col + str(hrows + 8759) + ',">0",Detail!C' + str(hrows) + ':Detail!C' \
            + str(hrows + 8759) + ')-C' + str(ss_row + 1)
        ss.cell(row=ss_row, column=3).number_format = '#,##0'
        ss.cell(row=ss_row, column=4).value = '=C' + str(ss_row) + '/C' + str(ss_row + 2)
        ss.cell(row=ss_row, column=4).number_format = '#,##0.0%'
        ss_row += 1
        ss.cell(row=ss_row, column=1).value = 'Shortfall'
        ss.cell(row=ss_row, column=3).value = '=SUMIF(Detail!' + last_col + str(hrows) + ':Detail!' \
            + last_col + str(hrows + 8759) + ',">0",Detail!' + last_col + str(hrows) + ':Detail!' \
            + last_col + str(hrows + 8759) + ')'
        ss.cell(row=ss_row, column=3).number_format = '#,##0'
        ss.cell(row=ss_row, column=4).value = '=C' + str(ss_row) + '/C' + str(ss_row + 1)
        ss.cell(row=ss_row, column=4).number_format = '#,##0.0%'
        ss_row += 1
        ss.cell(row=ss_row, column=1).value = 'Total Load'
        ss.cell(row=ss_row, column=1).font = bold
        ss.cell(row=ss_row, column=3).value = '=SUM(C' + str(ss_row - 2) + ':C' + str(ss_row - 1) + ')'
        ss.cell(row=ss_row, column=3).number_format = '#,##0'
        ss.cell(row=ss_row, column=3).font = bold
        ss_row += 1
        ss.cell(row=ss_row, column=1).value = 'Surplus'
        ss.cell(row=ss_row, column=3).value = '=-SUMIF(Detail!' + last_col + str(hrows) + ':Detail!' \
            + last_col + str(hrows + 8759) + ',"<0",Detail!' + last_col + str(hrows) \
            + ':Detail!' + last_col + str(hrows + 8759) + ')'
        ss.cell(row=ss_row, column=3).number_format = '#,##0'
        ss.cell(row=ss_row, column=4).value = '=C' + str(ss_row) + '/C' + str(ss_row - 1)
        ss.cell(row=ss_row, column=4).number_format = '#,##0.0%'
        ss_row += 2
        ss.cell(row=ss_row, column=1).value = 'Data sources:'
        ss.cell(row=ss_row, column=1).font = bold
        ss_row += 1
        ss.cell(row=ss_row, column=1).value = 'Scenarios folder'
        ss.cell(row=ss_row, column=2).value = self.scenarios
        ss.merge_cells('B' + str(ss_row) + ':G' + str(ss_row))
        ss_row += 1
        ss.cell(row=ss_row, column=1).value = 'Powermatch data file'
        if pm_data_file[: len(self.scenarios)] == self.scenarios:
            pm_data_file = pm_data_file[len(self.scenarios):]
        ss.cell(row=ss_row, column=2).value = pm_data_file
        ss.merge_cells('B' + str(ss_row) + ':G' + str(ss_row))
        ss_row += 1
        ss.cell(row=ss_row, column=1).value = 'Constraints worksheet'
        ss.cell(row=ss_row, column=2).value = str(self.files[C].text()) \
               + '.' + str(self.sheets[C].currentText())
        ss.merge_cells('B' + str(ss_row) + ':G' + str(ss_row))
        ss_row += 1
        ss.cell(row=ss_row, column=1).value = 'Facility worksheet'
        ss.cell(row=ss_row, column=2).value = str(self.files[G].text()) \
               + '.' + str(self.sheets[G].currentText())
        ss.merge_cells('B' + str(ss_row) + ':G' + str(ss_row))
        self.progressbar.setValue(7)
        try:
            if self.adjustby is not None:
                adjusted = ''
                for key, value in iter(sorted(self.adjustby.items())):
                    if value != 1:
                        adjusted += key + ': ' + str(value) + '; '
                if len(adjusted) > 0:
                    ss_row += 1
                    ss.cell(row=ss_row, column=1).value = 'Renewable inputs adjusted'
                    ss.cell(row=ss_row, column=2).value = adjusted[:-2]
        except:
            pass
        ss.freeze_panes = 'B4'
        ss.activeCell = 'B4'
        ds.save(data_file)
        self.progressbar.setValue(10)
        j = data_file.rfind('/')
        data_file = data_file[j + 1:]
        msg = '%s created (%.2f seconds)' % (data_file, time.time() - start_time)
        msg = '%s created.' % data_file
        self.log.setText(msg)
        self.progressbar.setHidden(True)
        self.progressbar.setValue(0)

    def show_ProgressBar(self, maximum, msg, title):
        if self.opt_progressbar is None:
            self.opt_progressbar = ProgressBar(maximum=maximum, msg=msg, title=title)
            self.opt_progressbar.setWindowModality(QtCore.Qt.WindowModal)
            self.connect(self.opt_progressbar, QtCore.SIGNAL('progress'), self.opt_progressbar.progress)
            self.connect(self.opt_progressbar, QtCore.SIGNAL('range'), self.opt_progressbar.range)
            self.opt_progressbar.show()
            self.opt_progressbar.setVisible(False)
            self.activateWindow()
        else:
            self.opt_progressbar.range(0, maximum, msg=msg)

    def optClicked(self):

        def create_starting_population(individuals, chromosome_length):
            # Set up an initial array of all zeros
            population = np.zeros((individuals, chromosome_length))
            # Loop through each row (individual)
            for i in range(individuals):
                # Choose a random number of ones to create
                ones = random.randint(0, chromosome_length)
                # Change the required number of zeros to ones
                population[i, 0:ones] = 1
                # Sfuffle row
                np.random.shuffle(population[i])

            return population

        def select_individual_by_tournament(population, scores):
            # Get population size
            population_size = len(scores)

            # Pick individuals for tournament
            fighter_1 = random.randint(0, population_size-1)
            fighter_2 = random.randint(0, population_size-1)

            # Get fitness score for each
            fighter_1_fitness = scores[fighter_1]
            fighter_2_fitness = scores[fighter_2]

            # Identify undividual with lowest LCOE
            # Fighter 1 will win if score are equal
            if fighter_1_fitness <= fighter_2_fitness:
                winner = fighter_1
            else:
                winner = fighter_2

            # Return the chromsome of the winner
            return population[winner, :]

        def breed_by_crossover(parent_1, parent_2):
            # Get length of chromosome
            chromosome_length = len(parent_1)

            # Pick crossover point, avoding ends of chromsome
            crossover_point = random.randint(1,chromosome_length-1)

            # Create children. np.hstack joins two arrays
            child_1 = np.hstack((parent_1[0:crossover_point],
                                parent_2[crossover_point:]))

            child_2 = np.hstack((parent_2[0:crossover_point],
                                parent_1[crossover_point:]))

            # Return children
            return child_1, child_2

        def calculate_fitness(population):
            fitness_scores = []
            for chromosome in population:
                op_data = []
                shortfall = op_load[:]
                # now get random amount of generation per technology (both RE and non-RE)
                for gen, value in dict_order.items():
                    capacity = dict_order[gen][3]
                    for c in range(value[1], value[2]):
                        if chromosome[c]:
                            capacity = capacity + capacities[c]
                    self.adjustby[gen] = capacity / dict_order[gen][0]
                    if gen in orig_tech.keys():
                        op_data.append([gen, capacity, 0., '', '', '', ''])
                        adjust = capacity / dict_order[gen][0]
                        for h in range(len(shortfall)):
                            g = orig_tech[gen][h] * adjust
                            op_data[-1][2] += g
                            shortfall[h] = shortfall[h] - g
                        continue
                    if generators[gen].constraint in constraints and \
                      constraints[generators[gen].constraint].category == 'Storage': # storage
                        storage = [0., 0., 0., 0.] # capacity, initial, min level, max drain
                        storage[0] = capacity
                        storage[1] = generators[gen].initial
                        if constraints[generators[gen].constraint].capacity_min > 0:
                            storage[2] = capacity * constraints[generators[gen].constraint].capacity_min
                        if constraints[generators[gen].constraint].capacity_max > 0:
                            storage[3] = capacity * constraints[generators[gen].constraint].capacity_max
                        else:
                            storage[3] = capacity
                        recharge = [0., 0.] # cap, loss
                        if constraints[generators[gen].constraint].recharge_max > 0:
                            recharge[0] = capacity * constraints[generators[gen].constraint].recharge_max
                        else:
                            recharge[0] = capacity
                        if constraints[generators[gen].constraint].recharge_loss > 0:
                            recharge[1] = constraints[generators[gen].constraint].recharge_loss
                        discharge = [0., 0.] # cap, loss
                        if constraints[generators[gen].constraint].discharge_max > 0:
                            discharge[0] = capacity * constraints[generators[gen].constraint].discharge_max
                        if constraints[generators[gen].constraint].discharge_loss > 0:
                            discharge[1] = constraints[generators[gen].constraint].discharge_loss
                        if constraints[generators[gen].constraint].parasitic_loss > 0:
                            parasite = constraints[generators[gen].constraint].parasitic_loss / 24.
                        else:
                            parasite = 0.
                        storage_carry = storage[1] # generators[gen].initial
                        storage_bal = []
                        storage_can = 0.
                        for row in range(8760):
                            if storage_carry > 0:
                                storage_carry = storage_carry * (1 - parasite)
                            storage_loss = 0.
                            if shortfall[row] < 0:  # excess generation
                                can_use = - (storage[0] - storage_carry) * (1 / (1 - recharge[1]))
                                if can_use < 0: # can use some
                                    if shortfall[row] > can_use:
                                        can_use = shortfall[row]
                                    if can_use < - recharge[0] * (1 / (1 - recharge[1])):
                                        can_use = - recharge[0]
                                else:
                                    can_use = 0.
                                storage_carry -= (can_use * (1 - recharge[1]))
                                shortfall[row] -= can_use
                            else: # shortfall
                                can_use = shortfall[row] * (1 / (1 - discharge[1]))
                                if can_use > storage_carry - storage[2]:
                                    can_use = storage_carry - storage[2]
                                if can_use > 0:
                                    storage_loss = can_use * discharge[1]
                                    storage_carry -= can_use
                                    shortfall[row] -= (can_use - storage_loss)
                                    if storage_carry < 0:
                                        storage_carry = 0
                                else:
                                    can_use = 0.
                            storage_bal.append(storage_carry)
                            if can_use > 0:
                                storage_can += can_use
                        op_data.append([gen, storage[0], storage_can, '', '', '', ''])
                    else:
                        gen_can = 0.
                        for row in range(8760):
                            if shortfall[row] >= 0: # shortfall?
                                if shortfall[row] >= capacity:
                                    shortfall[row] = shortfall[row] - capacity
                                    gen_can += capacity
                                else:
                                    gen_can += shortfall[row]
                                    shortfall[row] = 0
                        op_data.append([gen, capacity, gen_can, '', '', '', ''])
                cap_sum = 0.
                gen_sum = 0.
                re_sum = 0.
                cost_sum = 0.
                co2_sum = 0.
                for sp in range(len(op_data)):
                    if op_data[sp][1] > 0:
                        cap_sum += op_data[sp][1]
                        op_data[sp][3] = op_data[sp][2] / op_data[sp][1] / 8760
                    gen_sum += op_data[sp][2]
                    gen = op_data[sp][0]
                    if gen in orig_tech.keys():
                        re_sum += op_data[sp][2]
                    if gen not in generators:
                        continue
                    if generators[gen].lcoe > 0:
                        op_data[sp][4] = generators[gen].lcoe * generators[gen].lcoe_cf * 8760 * op_data[sp][1]
                        if op_data[sp][1] > 0 and op_data[sp][3] > 0:
                            op_data[sp][5] = op_data[sp][4] / 8760 / op_data[sp][3] / op_data[sp][1]
                        cost_sum += op_data[sp][4]
                    if generators[gen].emissions > 0:
                        op_data[sp][6] = op_data[sp][2] * generators[gen].emissions
                        co2_sum += op_data[sp][6]
                if cap_sum > 0:
                    cs = gen_sum / cap_sum / 8760
                else:
                    cs = ''
                if gen_sum > 0:
                    gs = cost_sum / gen_sum
                else:
                    gs = ''
                sf_sums = [0., 0., 0.]
                for sf in range(len(shortfall)):
                    if shortfall[sf] > 0:
                        sf_sums[0] += shortfall[sf]
                        sf_sums[2] += op_load[sf]
                    else:
                        sf_sums[1] += shortfall[sf]
                        sf_sums[2] += op_load[sf]
                if (sf_sums[2] - sf_sums[0]) / op_load_tot < 1:
                    fitness_scores.append(200)
                else:
                    fitness_scores.append(gs)
            if len(population) == 1:
                op_data.append(['Total', cap_sum, gen_sum, cs, cost_sum, gs, co2_sum])
                if self.carbon_price > 0:
                    cc = co2_sum * self.carbon_price
                    cs = (cost_sum + cc) / gen_sum
                    if self.carbon_price == int(self.carbon_price):
                       cp = str(int(self.carbon_price))
                    else:
                       cp = '{:.2f}'.format(self.carbon_price)
                    op_data.append(['Carbon Cost ($' + cp + ')', '', '', '', cc, cs])
                op_data.append(['RE %age', round(re_sum * 100. / gen_sum, 1)])
                op_data.append(' ')
                op_data.append('Load Analysis')
                pct = '{:.1%})'.format((sf_sums[2] - sf_sums[0]) / op_load_tot)
                op_data.append(['Load met (' + pct, '', sf_sums[2] - sf_sums[0],])
                pct = '{:.1%})'.format(sf_sums[0] / op_load_tot)
                op_data.append(['Shortfall (' + pct, '', sf_sums[0]])
                op_data.append(['Total Load', '', op_load_tot])
                pct = '{:.1%})'.format( -sf_sums[1] / op_load_tot)
                op_data.append(['Surplus (' + pct, '', -sf_sums[1]])
                return op_data
            else:
                return fitness_scores

        def optQuitClicked(event):
            self.optExit = True
            optDialog.close()

        self.optExit = False
        self.log.setText('Optimise processing started')
        details = self.details
        try:
            if str(self.files[C].text()).find('/') >= 0:
                ts = xlrd.open_workbook(str(self.files[C].text()))
            else:
                ts = xlrd.open_workbook(self.scenarios + str(self.files[C].text()))
            ws = ts.sheet_by_name(self.sheets[C].currentText())
            constraints = self.getConstraints(ws)
            ts.release_resources()
            del ts
        except:
            self.log.setText('Error accessing Constraints.')
            return
        try:
            if str(self.files[G].text()).find('/') >= 0:
                ts = xlrd.open_workbook(str(self.files[G].text()))
            else:
                ts = xlrd.open_workbook(self.scenarios + str(self.files[G].text()))
            ws = ts.sheet_by_name(self.sheets[G].currentText())
            generators = self.getGenerators(ws)
            ts.release_resources()
            del ts
        except:
            self.log.setText('Error accessing Generators.')
            self.progressbar.setHidden(True)
            return
        try:
            if str(self.files[O].text()).find('/') >= 0:
                ts = xlrd.open_workbook(str(self.files[O].text()))
            else:
                ts = xlrd.open_workbook(self.scenarios + str(self.files[O].text()))
            ws = ts.sheet_by_name(self.sheets[O].currentText())
            optimisation = self.getoptimisation(ws, generators)
            ts.release_resources()
            del ts
        except:
            self.log.setText('Error accessing Optimisation.')
            self.progressbar.setHidden(True)
            return
        if str(self.files[D].text()).find('/') >= 0:
            pm_data_file = str(self.files[D].text())
        else:
            pm_data_file = self.scenarios + str(self.files[D].text())
        if pm_data_file[-4:] == '.xls': #xls format
            self.log.setText('Not an option for Powerbalance')
            return
        ts = oxl.load_workbook(pm_data_file)
        ws = ts.active
        top_row = ws.max_row - 8760
        if ws.cell(row=top_row, column=1).value != 'Hour' or ws.cell(row=top_row, column=2).value != 'Period':
            self.log.setText('not a Powermatch data spreadsheet')
            self.progressbar.setHidden(True)
            return
        typ_row = top_row - 1
        while typ_row > 0:
            if ws.cell(row=typ_row, column=3).value in tech_names:
                break
            typ_row -= 1
        else:
            self.log.setText('no suitable data')
            return
        icap_row = typ_row + 1
        while icap_row < top_row:
            if ws.cell(row=icap_row, column=1).value == 'Capacity (MW)':
                break
            icap_row += 1
        else:
            self.log.setText('no capacity data')
            return
        optExit = False
        optDialog = QtGui.QDialog()
        grid = QtGui.QGridLayout()
        grid.addWidget(QtGui.QLabel('Adjust load'), 0, 0)
        optLoad = QtGui.QDoubleSpinBox()
        optLoad.setRange(-1, 25)
        optLoad.setDecimals(3)
        optLoad.setSingleStep(.1)
        optLoad.setValue(1)
        grid.addWidget(optLoad, 0, 1)
        grid.addWidget(QtGui.QLabel('Multiplier for input Load'), 0, 2)
        grid.addWidget(QtGui.QLabel('Population size'), 1, 0)
        optPopn = QtGui.QSpinBox()
        optPopn.setRange(10, 500)
        optPopn.setSingleStep(10)
        optPopn.setValue(self.optimise_population)
        optPopn.valueChanged.connect(self.changes)
        grid.addWidget(optPopn, 1, 1)
        grid.addWidget(QtGui.QLabel('Size of population'), 1, 2)
        grid.addWidget(QtGui.QLabel('No. of generations'), 2, 0)
        optGenn = QtGui.QSpinBox()
        optGenn.setRange(10, 200)
        optGenn.setSingleStep(10)
        optGenn.setValue(self.optimise_generations)
        optGenn.valueChanged.connect(self.changes)
        grid.addWidget(optGenn, 2, 1)
        grid.addWidget(QtGui.QLabel('Number of generations (iterations)'), 2, 2)
        grid.addWidget(QtGui.QLabel('Exit if stable'), 3, 0)
        optStop = QtGui.QSpinBox()
        optStop.setRange(0, 50)
        optStop.setSingleStep(10)
        optStop.setValue(self.optimise_stop)
        optStop.valueChanged.connect(self.changes)
        grid.addWidget(optStop, 3, 1)
        grid.addWidget(QtGui.QLabel('Exit if LCOE remains the same after this many iterations'), 3, 2)
        quit = QtGui.QPushButton('Quit', self)
        grid.addWidget(quit, 4, 0)
        quit.clicked.connect(optQuitClicked)
        show = QtGui.QPushButton('Proceed', self)
        grid.addWidget(show, 4, 1)
        show.clicked.connect(optDialog.close)
        optDialog.setLayout(grid)
        optDialog.setWindowTitle('Choose Optimisation Parameters')
        optDialog.setWindowIcon(QtGui.QIcon('sen_icon32.ico'))
        optDialog.exec_()
        if self.optExit: # a fudge to exit
            self.log.setText('Execution aborted.')
            return
        self.adjust_re = True
        re_capacities = []
        re_capacity = 0.
        orig_load = []
        load_col = -1
        orig_tech = {}
        dict_order = {} # rely on will be processed in added order
        # first get original renewables generation from data sheet
        for col in range(3, ws.max_column + 1):
            try:
                valu = ws.cell(row=typ_row, column=col).value.replace('-','')
                i = tech_names.index(valu)
            except:
                break # ?? or continue??
            if tech_names[i] != 'Load':
                try:
                    if ws.cell(row=icap_row, column=col).value <= 0:
                        continue
                except:
                    continue
                dict_order[tech_names[i]] = [ws.cell(row=icap_row, column=col).value, 0, 0, 0]
            orig_tech[tech_names[i]] = []
            for row in range(top_row + 1, ws.max_row + 1):
                orig_tech[tech_names[i]].append(ws.cell(row=row, column=col).value)
        ts.close()
        # now add scheduled generation
        for itm in range(self.order.count()):
            tech = str(self.order.item(itm).text())
            if tech not in dict_order.keys():
                dict_order[tech] = [generators[tech].capacity, 0, 0, 0]
        capacities = []
        for tech in dict_order.keys():
            dict_order[tech][1] = len(capacities) # first entry
            if optimisation[tech].approach == 'Discrete':
                capacities.extend(optimisation[tech].capacities)
                dict_order[tech][2] = len(capacities) # last entry
            elif optimisation[tech].approach == 'Range':
                ctr = int((optimisation[tech].capacity_max - optimisation[tech].capacity_min) / \
                          optimisation[tech].capacity_step)
                capacities.extend([optimisation[tech].capacity_step] * ctr)
                dict_order[tech][2] = len(capacities)
                dict_order[tech][3] = optimisation[tech].capacity_min
            else:
                dict_order[tech][2] = len(capacities)
        # chromosome = [1] * int(len(capacities) / 2) + [0] * (len(capacities) - int(len(capacities) / 2))
        # we have the original data - from here down we can do our multiple optimisations
        self.adjustby = {'Load': optLoad.value()}
        headers = ['Facility', 'Capacity (MW)', 'Subtotal (MWh)', 'CF', 'Cost ($)', 'LCOE ($/MWh)', 'Emissions (tCO2e)']
        op_data = []
        if self.adjustby['Load'] == 1:
            op_load = orig_tech['Load'][:]
        else:
            op_load = []
            for load in orig_tech['Load']:
                op_load.append(load * self.adjustby['Load'])
        op_load_tot = sum(op_load)
        # Set general parameters
        chromosome_length = len(capacities)
        self.optimise_population = optPopn.value()
        population_size = self.optimise_population
        self.optimise_generations = optGenn.value()
        maximum_generation = self.optimise_generations
        self.optimise_stop = optStop.value()
        best_score_progress = [] # Tracks progress
        target = 0. # aim for this LCOE
        scores = []
        self.show_ProgressBar(maximum=optGenn.value(), msg='Process generations', title='SIREN - Powermatch Progress')
        self.opt_progressbar.setVisible(True)
        start_time = time.time()
        # Create starting population
        population = create_starting_population(population_size, chromosome_length)
        # Display best score in starting population
        scores = calculate_fitness(population)
        best_score = np.min(scores)
        lowest_score = best_score
        best = scores.index(best_score)
        lowest_chrom = population[best]
        print ('Starting LCOE: ',best_score)
        lcoe_value = best_score
        lcoe_ctr = 1
        # Add starting best score to progress tracker
        best_score_progress.append(best_score)
        # Now we'll go through the generations of genetic algorithm
        for generation in range(maximum_generation):
            tim = (time.time() - start_time)
            if tim < 60:
                tim = ' (%.1f secs)' % tim
            else:
                tim = ' (%.2f mins)' % (tim / 60.)
            self.opt_progressbar.emit(QtCore.SIGNAL('progress'), generation,
                'Processing generation ' + str(generation) + tim)
            QtGui.qApp.processEvents()
            if not self.opt_progressbar.be_open:
                break
        # Create an empty list for new population
            new_population = []
            # Create new population generating two children at a time
            for i in range(int(population_size/2)):
                parent_1 = select_individual_by_tournament(population, scores)
                parent_2 = select_individual_by_tournament(population, scores)
                child_1, child_2 = breed_by_crossover(parent_1, parent_2)
                new_population.append(child_1)
                new_population.append(child_2)

            # Replace the old population with the new one
            population = np.array(new_population)

            # Score best solution, and add to tracker
            scores = calculate_fitness(population)
            best_score = np.min(scores)
            best_score_progress.append(best_score)
            if best_score < lowest_score:
                best = scores.index(best_score)
                lowest_score = best_score
                lowest_chrom = population[best]
            if self.optimise_stop > 0:
                if best_score == lcoe_value:
                    lcoe_ctr += 1
                    if lcoe_ctr >= self.optimise_stop:
                        break
                else:
                    lcoe_value = best_score
                    lcoe_ctr = 1
        self.opt_progressbar.setVisible(False)
        self.opt_progressbar.close()
        tim = (time.time() - start_time)
        if tim < 60:
            tim = ' (%.1f secs)' % tim
        else:
            tim = ' (%.2f mins)' % (tim / 60.)
        msg = 'Optimise completed' + tim
        best = scores.index(best_score)
        if best_score > lowest_score:
            msg += ' Try more generations.'
            op_data = calculate_fitness([lowest_chrom])
        else:
            op_data = calculate_fitness([population[best]])
        self.log.setText(msg)
        self.progressbar.setHidden(True)
        self.progressbar.setValue(0)
        # GA has completed required generation
        print ('End LCOE: ', best_score)
        # Plot progress
        rcParams['savefig.directory'] = self.scenarios
        plt.figure('optimise_lcoe')
        plt.title('Optimise LCOE using Genetic Algorithm')
        plt.plot(best_score_progress)
        plt.xlabel('Optimise Cycle (' + str(len(best_score_progress)) + ' generations)')
        plt.ylabel('Best LCOE ($/MWh)')
        plt.show()
        list(map(list, list(zip(*op_data))))
        op_pts = [0, 3, 0, 2, 0, 2, 0]
        dialog = displaytable.Table(op_data, title=self.sender().text(), fields=headers,
                 save_folder=self.scenarios, sortby='', decpts=op_pts)
        dialog.exec_()
        return

if "__main__" == __name__:
    app = QtGui.QApplication(sys.argv)
    ex = powerMatch()
    app.exec_()
    app.deleteLater()
    sys.exit()
